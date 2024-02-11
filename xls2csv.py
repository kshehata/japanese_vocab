from openpyxl import load_workbook
import csv
import sys

if len(sys.argv) < 2:
    print("Usage: xls2csv filenames")
    sys.exit(1)

def format_cell(cell):
    if cell.value is not None and cell.data_type == "n":
        return f'{int(cell.value):,}'
    return cell.value

for filename in sys.argv[1:]:
    wb = load_workbook(filename)
    for sheetname in wb.sheetnames:
        print(sheetname)
        ws = wb[sheetname]
        with open(sheetname + ".csv", "w") as csvfile:
            csv_writer = csv.writer(csvfile, lineterminator="\n")
            for row in ws.iter_rows():
                if all([cell.value is None for cell in row]): continue
                csv_writer.writerow([format_cell(cell) for cell in row])
